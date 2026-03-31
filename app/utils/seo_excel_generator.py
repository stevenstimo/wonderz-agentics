"""
SEO Plan Excel generator — Keyword Plan, Silo, Quick Wins, Strategie, Content Gaps,
GSC Performance, Markt Expansie, Interne Links, Concurrent Analyse, Legenda (openpyxl).
"""
import json
import logging
import os
import re
from collections import defaultdict
from datetime import datetime
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.utils.seo_classification import (
    cap_content_gap_priority,
    quick_win_score,
)

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HIGH_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
MEDIUM_FILL = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
YELLOW_NOTE_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
DATA_FONT = Font(name="Arial", size=10)

STRATEGY_CONTENT_PREAMBLE = """## Content doelstelling

[Door de gebruiker in te vullen — voorbeeld:]
\"Bezoekers die twijfelen over [onderwerp] overtuigen dat [merknaam] de betrouwbare partner is.\"

Leidende vraag per artikel: Wat moet de lezer begrijpen, anders zien of voelen na het lezen?"""

BUSINESS_RELEVANCE_COMMENT = (
    "Door agent ingevuld op basis van keyword-inhoud. Vereist handmatige review door de strateeg."
)

GSC_LABEL_FILLS = {
    "✅ Sterk": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    "🟡 Optimaliseer": PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid"),
    "🟠 Aanpakken": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
    "🔴 Zwak": PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),
    "⬜ Ontbreekt": PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
}

# Kolomvolgorde v2 — Markt … URL
KEYWORD_PLAN_COLUMNS = [
    "Markt",
    "Taal",
    "Keyword",
    "Cluster",
    "Silo",
    "Doelgroep",
    "Volume",
    "Trend",
    "KD",
    "CPC",
    "Click Potential",
    "Intent",
    "SERP Features",
    "Status (GSC)",
    "Positie (GSC)",
    "Clicks (GSC)",
    "Impressies (GSC)",
    "CTR (GSC)",
    "Content Type",
    "Prioriteit",
    "Business Relevance",
    "Status",
    "Week",
    "URL",
]

CONTENT_GAPS_COLUMNS = [
    "Markt",
    "Keyword",
    "Volume",
    "KD",
    "Intent",
    "SERP Features",
    "Silo",
    "Content Type",
    "Prioriteit",
]


def _taal_for_market(market: Any) -> str:
    m = str(market or "NL").strip().upper()
    if m == "UK":
        return "en"
    if m == "DE":
        return "de"
    return "nl"


def _ensure_output_dir() -> str:
    out_dir = "/tmp/seo_plans"
    os.makedirs(out_dir, exist_ok=True)
    return out_dir


def _row_fill_for_priority(priority: str) -> Optional[PatternFill]:
    if priority == "HIGH":
        return HIGH_FILL
    if priority == "MEDIUM":
        return MEDIUM_FILL
    return None


def _apply_header_style(ws, row_num: int = 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def _auto_column_width(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 15
        for row in range(1, min(ws.max_row + 1, 100)):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                max_len = max(max_len, min(len(str(val)), 50))
        ws.column_dimensions[get_column_letter(col)].width = max_len


def _build_silo_overview(keywords: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    silos: Dict[str, Dict[str, Any]] = {}
    for k in keywords:
        silo = k.get("silo") or ""
        if not silo:
            continue
        if silo not in silos:
            silos[silo] = {"name": silo, "count": 0, "volume": 0, "kd_sum": 0, "content_types": set()}
        silos[silo]["count"] += 1
        silos[silo]["volume"] += k.get("volume") or 0
        silos[silo]["kd_sum"] += k.get("kd") or 0
        if k.get("content_type"):
            silos[silo]["content_types"].add(k["content_type"])

    result = []
    for s in silos.values():
        count = s["count"]
        avg_kd = s["kd_sum"] / count if count else 0
        ct = ", ".join(sorted(s["content_types"])) if s["content_types"] else "Blog"
        approach = f"Focus op {ct} content. Gemiddelde KD: {avg_kd:.0f}."
        result.append({
            "Silo naam": s["name"],
            "Aantal keywords": count,
            "Totaal volume": s["volume"],
            "Gemiddelde KD": round(avg_kd, 1),
            "Aanbevolen aanpak": approach,
        })
    return sorted(result, key=lambda x: -x["Totaal volume"])


def _format_numeric(value: Any, decimals: int = 2) -> Any:
    """Round numeric values before writing to Excel to avoid float artifacts (e.g. 0.15000000000000002)."""
    if value is None:
        return None
    if isinstance(value, str) and not str(value).strip():
        return None
    try:
        n = float(value)
    except (TypeError, ValueError):
        return value
    r = round(n, decimals)
    if decimals == 0:
        return int(r)
    return r


def _ctr_display(ctr: Any) -> Any:
    if ctr is None:
        return None
    try:
        n = float(ctr)
        n = round(n, 6)
        if n == 0:
            return "0.0%"
        return f"{n * 100:.1f}%"
    except (TypeError, ValueError):
        return None


def _keyword_plan_row_data(k: Dict[str, Any]) -> List[Any]:
    gsc_pos = k.get("gsc_position")
    gsc_clicks = k.get("gsc_clicks")
    gsc_imp = k.get("gsc_impressions")
    gsc_ctr = k.get("gsc_ctr")
    pos_display = None
    if gsc_pos is not None and gsc_pos != 0:
        pos_display = _format_numeric(gsc_pos, 1)
    mkt = k.get("market") or "NL"
    cp = k.get("click_potential")
    vol = k.get("volume")
    kd = k.get("kd")
    cpc = k.get("cpc")
    return [
        mkt,
        _taal_for_market(mkt),
        k.get("keyword") or "",
        k.get("cluster") or "",
        k.get("silo") or "",
        k.get("audience_match") or "",
        _format_numeric(vol, 0) if vol is not None else None,
        k.get("trend") or "? Onbekend",
        _format_numeric(kd, 1) if kd is not None else None,
        _format_numeric(cpc, 2) if cpc is not None else None,
        _format_numeric(cp, 0) if cp is not None else None,
        k.get("intent") or "",
        k.get("serp_features") or "",
        k.get("gsc_label") or None,
        pos_display,
        _format_numeric(gsc_clicks, 0) if gsc_clicks is not None else None,
        _format_numeric(gsc_imp, 0) if gsc_imp is not None else None,
        _ctr_display(gsc_ctr) if gsc_ctr is not None else None,
        k.get("content_type") or "",
        k.get("priority") or "",
        k.get("business_relevance") or "?",
        k.get("plan_status") or "",
        k.get("plan_week") or "",
        k.get("current_url") or "",
    ]


def _filter_quick_wins(keywords: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Status Aanpakken/Optimaliseer, volume≥100, KD≤40; sort by quick_win_score DESC."""
    result = []
    for k in keywords:
        label = k.get("gsc_label") or ""
        if label not in ("🟠 Aanpakken", "🟡 Optimaliseer"):
            continue
        if (k.get("volume") or 0) < 100:
            continue
        if (k.get("kd") or 100) > 40:
            continue
        pos = k.get("gsc_position")
        if pos is None:
            continue
        k2 = dict(k)
        k2["quick_win_score"] = quick_win_score(
            k.get("volume", 0),
            k.get("kd", 0),
            k.get("gsc_clicks"),
            float(pos) if pos is not None else None,
        )
        result.append(k2)
    result = sorted(result, key=lambda x: x.get("quick_win_score") or 0, reverse=True)
    return [k for k in result if (k.get("quick_win_score") or 0) > 0]


def _generate_strategy_metrics(
    keywords: List[Dict[str, Any]],
    silos: List[Dict[str, Any]],
    gsc_connected: bool,
) -> str:
    high = sum(1 for k in keywords if k.get("priority") == "HIGH")
    med = sum(1 for k in keywords if k.get("priority") == "MEDIUM")
    low = sum(1 for k in keywords if k.get("priority") == "LOW")
    total_vol = sum(k.get("volume") or 0 for k in keywords)
    nl = sum(1 for k in keywords if (k.get("market") or "NL").upper() == "NL")
    uk = sum(1 for k in keywords if (k.get("market") or "").upper() == "UK")
    de = sum(1 for k in keywords if (k.get("market") or "").upper() == "DE")
    lines = [
        "# Strategie — metrics",
        "",
        f"- Totaal keywords: {len(keywords)} | Totaal zoekvolume (som): {total_vol}",
        f"- Prioriteit: HIGH {high} | MEDIUM {med} | LOW {low}",
        f"- Markten in plan: NL {nl} | UK {uk} | DE {de}",
        f"- GSC-data gekoppeld: {'ja' if gsc_connected else 'nee (upload zonder client/GSC of geen token)'}",
        "",
        "## Focus silo's (top 5 op volume)",
    ]
    for s in silos[:5]:
        lines.append(
            f"- {s.get('Silo naam', '')}: {s.get('Aantal keywords', 0)} keywords, "
            f"volume {s.get('Totaal volume', 0)}, gem. KD {s.get('Gemiddelde KD', 0)}"
        )
    lines.extend([
        "",
        "## Aanbevolen volgorde",
        "1. Quick wins (🟡/🟠 in GSC, volume ≥100, KD ≤40)",
        "2. HIGH-prioriteit met haalbare KD",
        "3. MEDIUM clusters en uitbreiding UK/DE (zie tab Markt Expansie)",
    ])
    return "\n".join(lines)


def _write_gsc_performance_sheet(wb: openpyxl.Workbook, perf: Optional[Dict[str, Any]], site_url: Optional[str]) -> None:
    ws = wb.create_sheet("📊 GSC Performance")
    row = 1
    ws.cell(row=row, column=1, value="Google Search Console — overzicht (laatste periode)")
    row += 1
    if site_url:
        ws.cell(row=row, column=1, value=f"Property: {site_url}")
        row += 1
    if not perf:
        ws.cell(row=row, column=1, value="Geen GSC-data beschikbaar voor dit plan.")
        _auto_column_width(ws)
        return

    ws.cell(row=row, column=1, value=f"Periode: {perf.get('date_start')} — {perf.get('date_end')}")
    row += 2

    def _table(title: str, headers: List[str], rows: List[Dict[str, Any]], keys: List[str]):
        nonlocal row
        ws.cell(row=row, column=1, value=title)
        row += 1
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        _apply_header_style(ws, row_num=row)
        row += 1
        for rdata in rows:
            for c, key in enumerate(keys, 1):
                ws.cell(row=row, column=c, value=rdata.get(key))
            row += 1
        row += 1

    tq = perf.get("top_queries") or []
    _table(
        "Top queries",
        ["Query", "Clicks", "Impressies", "CTR", "Positie"],
        tq,
        ["query", "clicks", "impressions", "ctr", "position"],
    )
    tp = perf.get("top_pages") or []
    _table(
        "Top pagina's",
        ["URL", "Clicks", "Impressies", "CTR", "Positie"],
        tp,
        ["page", "clicks", "impressions", "ctr", "position"],
    )
    tc = perf.get("countries") or []
    _table(
        "Landen",
        ["Land", "Clicks", "Impressies", "CTR", "Positie"],
        tc,
        ["country", "clicks", "impressions", "ctr", "position"],
    )
    ts = perf.get("timeseries") or []
    ws.cell(row=row, column=1, value="Trend (per dag)")
    row += 1
    for c, h in enumerate(["Datum", "Clicks", "Impressies"], 1):
        ws.cell(row=row, column=c, value=h)
    _apply_header_style(ws, row_num=row)
    row += 1
    for t in ts[:90]:
        ws.cell(row=row, column=1, value=t.get("date"))
        ws.cell(row=row, column=2, value=t.get("clicks"))
        ws.cell(row=row, column=3, value=t.get("impressions"))
        row += 1
    _auto_column_width(ws)


def _page_label_for_link(k: Dict[str, Any]) -> str:
    url = (k.get("current_url") or "").strip()
    if url:
        return url
    return (k.get("keyword") or "").strip() or "—"


def _pick_pillar_page(keywords_in_silo: List[Dict[str, Any]]) -> Dict[str, Any]:
    for x in keywords_in_silo:
        if (x.get("content_type") or "").strip().lower() == "pillar page":
            return x
    return max(keywords_in_silo, key=lambda x: int(x.get("volume") or 0))


def _normalize_competitor_token(part: str) -> str:
    p = part.strip()
    if not p:
        return ""
    p = re.sub(r"^https?://", "", p, flags=re.I)
    p = p.split("/")[0].strip()
    return p[:120]


def _competitor_tokens_from_cell(cell: Optional[str]) -> List[str]:
    if not cell or not str(cell).strip():
        return []
    s = str(cell).strip()
    if s.startswith("{") and "}" in s:
        try:
            obj = json.loads(s)
            if isinstance(obj, dict):
                out: List[str] = []
                for k in obj.keys():
                    t = _normalize_competitor_token(str(k).strip())
                    if t:
                        out.append(t)
                return out
        except (json.JSONDecodeError, TypeError, ValueError):
            pass
    parts = re.split(r"[,;\n|]+", s)
    out = []
    for p in parts:
        t = _normalize_competitor_token(p)
        if t:
            out.append(t)
    return out


def _cell_mentions_competitor(cell: Optional[str], token: str) -> bool:
    if not token:
        return False
    return token.lower() in (cell or "").lower()


def _competitor_approach_text(shared: int, silo_n: int) -> str:
    thresh = max(4, silo_n // 3) if silo_n else 4
    if shared >= thresh:
        return "Differentiëren"
    if shared >= 2:
        return "Aanvallen"
    return "Volgen"


def _write_internal_links_sheet(wb: openpyxl.Workbook, keywords: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet("Interne Links")
    headers = ["Pagina (van)", "Pagina (naar)", "Ankertekst", "Relatie", "Prioriteit"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _apply_header_style(ws)
    by_silo: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for k in keywords:
        silo = (k.get("silo") or "").strip() or "—"
        by_silo[silo].append(k)
    row = 2
    for _silo, items in by_silo.items():
        if len(items) < 2:
            continue
        pillar = _pick_pillar_page(items)
        p_lab = _page_label_for_link(pillar)
        p_kw = (pillar.get("keyword") or "").strip()
        for cl in items:
            if cl is pillar:
                continue
            c_lab = _page_label_for_link(cl)
            c_kw = (cl.get("keyword") or "").strip()
            pri = cl.get("priority") or "MEDIUM"
            ws.cell(row=row, column=1, value=p_lab)
            ws.cell(row=row, column=2, value=c_lab)
            ws.cell(row=row, column=3, value=c_kw or None)
            ws.cell(row=row, column=4, value="Pillar → Cluster")
            ws.cell(row=row, column=5, value=pri)
            for c in range(1, 6):
                ws.cell(row=row, column=c).font = DATA_FONT
            row += 1
            ws.cell(row=row, column=1, value=c_lab)
            ws.cell(row=row, column=2, value=p_lab)
            ws.cell(row=row, column=3, value=p_kw or None)
            ws.cell(row=row, column=4, value="Cluster → Pillar")
            ws.cell(row=row, column=5, value=pri)
            for c in range(1, 6):
                ws.cell(row=row, column=c).font = DATA_FONT
            row += 1
    if row == 2:
        ws.cell(row=2, column=1, value="Geen silo met 2+ keywords — geen interne linkkaart gegenereerd.")
    _auto_column_width(ws)


def _write_competitor_analysis_sheet(wb: openpyxl.Workbook, keywords: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet("Concurrent Analyse")
    headers = [
        "Silo",
        "Concurrent domein",
        "Gedeelde keywords",
        "Exclusieve concurrent keywords",
        "Gemiddelde KD gap",
        "Aanbevolen aanpak",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _apply_header_style(ws)

    has_cols = any((k.get("competitors") or "").strip() for k in keywords)
    if not has_cols:
        ws.cell(
            row=2,
            column=1,
            value="Voeg een Semrush-export toe met kolommen 'Competitors' en 'Content references' om deze sheet te vullen.",
        )
        ws.cell(
            row=3,
            column=1,
            value="Export uit Semrush (o.a. Keyword Magic) met competitor- en content reference-kolommen; upload als CSV of XLSX.",
        )
        _auto_column_width(ws)
        return

    by_silo: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for k in keywords:
        silo = (k.get("silo") or "").strip() or "—"
        by_silo[silo].append(k)

    row = 2
    for silo, skws in by_silo.items():
        all_tokens: set[str] = set()
        for k in skws:
            for t in _competitor_tokens_from_cell(k.get("competitors")):
                all_tokens.add(t)
        if not all_tokens:
            continue
        n_silo = len(skws)
        kd_vals = [float(x.get("kd") or 0) for x in skws]
        avg_silo_kd = sum(kd_vals) / len(kd_vals) if kd_vals else 0.0

        for token in sorted(all_tokens):
            shared = sum(1 for k in skws if _cell_mentions_competitor(k.get("competitors"), token))
            exclusive = max(0, n_silo - shared)
            matched = [k for k in skws if _cell_mentions_competitor(k.get("competitors"), token)]
            kd_m = [float(k.get("kd") or 0) for k in matched]
            avg_m = sum(kd_m) / len(kd_m) if kd_m else 0.0
            gap_val: Any = ""
            if matched:
                gap_val = round(avg_m - avg_silo_kd, 1)
            approach = _competitor_approach_text(shared, n_silo)

            ws.cell(row=row, column=1, value=silo)
            ws.cell(row=row, column=2, value=token)
            ws.cell(row=row, column=3, value=shared)
            ws.cell(row=row, column=4, value=exclusive)
            ws.cell(row=row, column=5, value=gap_val)
            ws.cell(row=row, column=6, value=approach)
            for c in range(1, 7):
                ws.cell(row=row, column=c).font = DATA_FONT
            row += 1

    if row == 2:
        ws.cell(row=2, column=1, value="Geen concurrent-domeinen gevonden in de kolom Competitors.")
    _auto_column_width(ws)


def _write_legenda_sheet(wb: openpyxl.Workbook, brand_name: str) -> None:
    ws = wb.create_sheet("Legenda")
    lines = [
        "Legenda — SEO Keyword Plan",
        f"Merken / plan: {brand_name or '—'}",
        "",
        "Trend: ↑ Stijgend / → Stabiel / ↓ Dalend / ? Onbekend (op basis van Semrush trend of maandvolumes).",
        "Business Relevance: HOOG / MEDIUM / LAAG / ? — agentsuggestie; altijd handmatig valideren.",
        "Click Potential: score 0–100; verwachte relatieve doorklik/CTR gegeven SERP-samenstelling en intent. "
        "Lager bij veel ads, featured snippet, AI Overview, shopping, knowledge panel, video. "
        "Transactioneel iets hoger; navigational iets lager.",
        "Status (kolom) & Week: suggesties door de agent (📋 Gepland + weekrange). "
        "Pas aan op capaciteit — zie ook Strategie-tab.",
        "URL: alleen gevuld wanneer de upload of keyword-export een URL bevat; anders leeg.",
        "",
        "Tabbladen Interne Links en Concurrent Analyse gebruiken silo-cluster en Semrush-kolommen Competitors / Content references.",
    ]
    for i, line in enumerate(lines, 1):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 100


def _write_markt_expansie_sheet(wb: openpyxl.Workbook, keywords: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet("🌍 Markt Expansie UK+DE")
    instructions = [
        "Markt Expansie — UK en DE",
        "",
        "1. Exporteer in Semrush Keyword Magic Tool aparte clusters voor United Kingdom en Germany.",
        "2. Zelfde kolomstructuur als de NL-export (o.a. Keyword, Volume, KD, Page/Cluster, SERP Features).",
        "3. Upload het NL-bestand verplicht; voeg optioneel UK- en DE-export toe in dezelfde run (web UI).",
        "4. Alle rijen krijgen een Markt-kolom (NL / UK / DE) voor filter in Excel.",
        "",
        "Keywords in dit plan per markt:",
    ]
    row = 1
    for line in instructions:
        ws.cell(row=row, column=1, value=line)
        row += 1
    counts: Dict[str, int] = {}
    for k in keywords:
        m = (k.get("market") or "NL").upper()
        counts[m] = counts.get(m, 0) + 1
    for m in ("NL", "UK", "DE"):
        ws.cell(row=row, column=1, value=f"{m}: {counts.get(m, 0)} keywords")
        row += 1
    _auto_column_width(ws)


def generate_seo_excel(
    keywords: List[Dict[str, Any]],
    brand_name: str,
    strategy_notes: Optional[str] = None,
    gsc_site_url: Optional[str] = None,
    gsc_performance: Optional[Dict[str, Any]] = None,
) -> str:
    """
    Multi-sheet Excel: Keyword Plan, Silo, Quick Wins, Strategie, Content Gaps,
    GSC Performance, Markt Expansie.
    """
    out_dir = _ensure_output_dir()
    safe_brand = "".join(c if c.isalnum() or c in " -_" else "_" for c in brand_name)[:30]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"SEO_Plan_{safe_brand}_{timestamp}.xlsx"
    filepath = os.path.join(out_dir, filename)

    wb = openpyxl.Workbook()
    gsc_connected = any(
        k.get("gsc_position") is not None or (k.get("gsc_clicks") is not None and k.get("gsc_clicks") != 0)
        for k in keywords
    )

    # Sheet 1: Keyword Plan
    ws1 = wb.active
    ws1.title = "Keyword Plan"
    note = (
        f"GSC keyword-match: laatste 90 dagen — {gsc_site_url} | Paginated query export"
        if gsc_site_url
        else "GSC kolommen: geen koppeling of geen data — Status ⬜ Ontbreekt"
    )
    ws1.cell(row=1, column=1, value=note)
    for col, header in enumerate(KEYWORD_PLAN_COLUMNS, 1):
        ws1.cell(row=2, column=col, value=header)
    _apply_header_style(ws1, row_num=2)
    br_col_idx = KEYWORD_PLAN_COLUMNS.index("Business Relevance") + 1
    ws1.cell(row=2, column=br_col_idx).comment = Comment(BUSINESS_RELEVANCE_COMMENT, "SEO Tool")

    for row_idx, k in enumerate(keywords, 3):
        row_data = _keyword_plan_row_data(k)
        gsc_label = k.get("gsc_label")
        for col_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            fill = GSC_LABEL_FILLS.get(gsc_label) if gsc_label else None
            if not fill:
                fill = _row_fill_for_priority(k.get("priority") or "")
            if fill:
                cell.fill = fill

    ws1.freeze_panes = "A3"
    ws1.auto_filter.ref = f"A2:{get_column_letter(ws1.max_column)}{ws1.max_row}"
    _auto_column_width(ws1)

    # Sheet 2: Silo Overzicht
    silos = _build_silo_overview(keywords)
    ws2 = wb.create_sheet("Silo Overzicht")
    silo_headers = ["Silo naam", "Aantal keywords", "Totaal volume", "Gemiddelde KD", "Aanbevolen aanpak"]
    for col, h in enumerate(silo_headers, 1):
        ws2.cell(row=1, column=col, value=h)
    _apply_header_style(ws2)
    for row_idx, s in enumerate(silos, 2):
        for col_idx, key in enumerate(silo_headers, 1):
            ws2.cell(row=row_idx, column=col_idx, value=s.get(key, ""))
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(silo_headers))}{ws2.max_row}"
    _auto_column_width(ws2)

    # Sheet 3: Quick Wins
    quick = _filter_quick_wins(keywords)
    quick_headers = KEYWORD_PLAN_COLUMNS + ["Quick Win Score"]
    ws3 = wb.create_sheet("Quick Wins")
    for col, h in enumerate(quick_headers, 1):
        ws3.cell(row=1, column=col, value=h)
    _apply_header_style(ws3)
    ws3.cell(row=1, column=br_col_idx).comment = Comment(BUSINESS_RELEVANCE_COMMENT, "SEO Tool")
    for row_idx, k in enumerate(quick, 2):
        score_val = _format_numeric(k.get("quick_win_score"), 1)
        row_data = _keyword_plan_row_data(k) + [score_val]
        gsc_label = k.get("gsc_label")
        for col_idx, val in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            fill = GSC_LABEL_FILLS.get(gsc_label) if gsc_label else None
            if not fill:
                fill = _row_fill_for_priority(k.get("priority") or "")
            if fill:
                cell.fill = fill
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(quick_headers))}{ws3.max_row}"
    _auto_column_width(ws3)

    # Sheet 4: Strategie
    notes = strategy_notes or _generate_strategy_metrics(keywords, silos, gsc_connected)
    preamble = STRATEGY_CONTENT_PREAMBLE.replace("[merknaam]", brand_name or "…")
    ws4 = wb.create_sheet("📝 Strategie")
    row4 = 1
    for line in preamble.split("\n"):
        ws4.cell(row=row4, column=1, value=line)
        row4 += 1
    row4 += 1
    note_cell = ws4.cell(row=row4, column=1, value="[Vul hier je contentdoelstelling in]")
    note_cell.fill = YELLOW_NOTE_FILL
    row4 += 1
    row4 += 1
    for line in notes.split("\n"):
        ws4.cell(row=row4, column=1, value=line)
        row4 += 1
    ws4.column_dimensions["A"].width = 90

    # Sheet 5: Content Gaps
    content_gaps = [k for k in keywords if k.get("gsc_label") == "⬜ Ontbreekt"]
    content_gaps_sorted = sorted(
        content_gaps,
        key=lambda x: (x.get("volume") or 0),
        reverse=True,
    )
    ws5 = wb.create_sheet("Content Gaps")
    ws5.cell(row=1, column=1, value="Content Gaps — geen GSC-ranking (⬜ Ontbreekt); prioriteit met authority-cap (KD>45 zonder ranking → max MEDIUM)")
    for col, h in enumerate(CONTENT_GAPS_COLUMNS, 1):
        ws5.cell(row=2, column=col, value=h)
    _apply_header_style(ws5, row_num=2)
    for row_idx, k in enumerate(content_gaps_sorted, 3):
        raw_p = k.get("priority") or "LOW"
        p = cap_content_gap_priority(
            raw_p,
            k.get("kd"),
            float(k["gsc_position"]) if k.get("gsc_position") is not None else None,
        )
        ws5.cell(row=row_idx, column=1, value=k.get("market") or "NL")
        ws5.cell(row=row_idx, column=2, value=k.get("keyword") or "")
        vol = k.get("volume")
        kd_gap = k.get("kd")
        ws5.cell(row=row_idx, column=3, value=_format_numeric(vol, 0) if vol is not None else None)
        ws5.cell(row=row_idx, column=4, value=_format_numeric(kd_gap, 1) if kd_gap is not None else None)
        ws5.cell(row=row_idx, column=5, value=k.get("intent") or "")
        ws5.cell(row=row_idx, column=6, value=k.get("serp_features") or "")
        ws5.cell(row=row_idx, column=7, value=k.get("silo") or "")
        ws5.cell(row=row_idx, column=8, value=k.get("content_type") or "")
        ws5.cell(row=row_idx, column=9, value=p)
        for col_idx in range(1, 10):
            cell = ws5.cell(row=row_idx, column=col_idx)
            cell.font = DATA_FONT
    ws5.freeze_panes = "A3"
    ws5.auto_filter.ref = f"A2:{get_column_letter(9)}{ws5.max_row}"
    _auto_column_width(ws5)

    _write_gsc_performance_sheet(wb, gsc_performance, gsc_site_url)
    _write_markt_expansie_sheet(wb, keywords)
    _write_internal_links_sheet(wb, keywords)
    _write_competitor_analysis_sheet(wb, keywords)
    _write_legenda_sheet(wb, brand_name)

    wb.save(filepath)
    logger.info("Generated SEO Excel: %s", filepath)
    return filepath
