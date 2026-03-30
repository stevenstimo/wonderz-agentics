"""
SEO Keyword CSV/XLSX/Numbers parser — parses Semrush/Ahrefs-style exports into keyword dicts.
Only keyword column is required; volume, kd, position, cpc, url, intent use aliases and defaults.
Supports Semrush Page (cluster), Click Potential, SERP Features, and Markt (NL/UK/DE).
"""
import csv
import io
import json
import logging
import os
import tempfile
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

COLUMN_ALIASES = {
    "keyword": [
        "keyword", "keywords", "query", "search query", "zoekwoord",
        "term", "search term", "key phrase", "keyphrase", "topic",
    ],
    "volume": [
        "search volume", "volume", "sv", "avg monthly searches",
        "monthly searches", "msv", "searches", "zoekvolume",
    ],
    "kd": [
        "keyword difficulty", "kd", "difficulty", "kd %", "competition",
        "keyword difficulty %", "moeilijkheid",
    ],
    "position": [
        "position", "pos", "rank", "ranking", "current position",
        "positie", "pos.", "current rank",
    ],
    "cpc": ["cpc", "cost per click", "cpc (usd)", "cpc (eur)"],
    "url": [
        "url", "current url", "landing page", "link", "content idea url",
    ],
    "cluster": [
        "page", "cluster", "keyword cluster", "topic cluster", "semrush page", "content page",
    ],
    "click_potential": [
        "click potential", "click_potential",
    ],
    "serp_features_raw": [
        "serp features", "serp feature", "serp_features",
    ],
    "intent": [
        "intent", "keyword intent", "intents", "keyword intents", "search intent",
    ],
}


def _normalize_col_key(col: str) -> str:
    return (col or "").strip().lower().replace(" ", "_").replace("-", "_")


def find_column(columns: List[str], aliases: List[str]) -> Optional[str]:
    normalized = {_normalize_col_key(col): col for col in columns}
    for alias in aliases:
        key = _normalize_col_key(alias)
        if key in normalized:
            return normalized[key]
    return None


def parse_csv(content: bytes) -> List[Dict[str, Any]]:
    """Parse CSV content into list of keyword dicts (original headers)."""
    text = content.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return []
    headers = [str(h).strip() for h in rows[0]]
    data = []
    for row in rows[1:]:
        if len(row) < len(headers):
            row = row + [""] * (len(headers) - len(row))
        elif len(row) > len(headers):
            row = row[: len(headers)]
        data.append(dict(zip(headers, row)))
    return data


def parse_xlsx(content: bytes) -> List[Dict[str, Any]]:
    """Parse XLSX content (first sheet) into list of keyword dicts."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = wb.active
    if not sheet:
        return []
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h or "").strip() for h in rows[0]]
    data = []
    for row in rows[1:]:
        row_values = [str(c) if c is not None else "" for c in row]
        if len(row_values) < len(headers):
            row_values = row_values + [""] * (len(headers) - len(row_values))
        elif len(row_values) > len(headers):
            row_values = row_values[: len(headers)]
        data.append(dict(zip(headers, row_values)))
    wb.close()
    return data


def parse_numbers_file(file_path: str) -> List[Dict[str, Any]]:
    """Parse Apple Numbers file (first sheet, first table) into list of dicts."""
    from numbers_parser import Document

    doc = Document(file_path)
    sheet = doc.sheets[0]
    table = sheet.tables[0]
    rows = list(table.iter_rows())
    if not rows:
        return []
    headers = [str(c.value) if c.value is not None else "" for c in rows[0]]
    data = []
    for row in rows[1:]:
        data.append(dict(zip(headers, [str(c.value) if c.value is not None else "" for c in row])))
    return data


def _parse_optional_float(val: Any) -> Optional[float]:
    if val is None or str(val).strip() == "":
        return None
    try:
        return float(str(val).replace(",", ".").replace("%", "").replace(" ", ""))
    except (ValueError, TypeError):
        return None


def _normalize_rows(rows: List[Dict[str, Any]], market: str = "NL") -> List[Dict[str, Any]]:
    """Map alias-based columns to standard keys; only keyword required; defaults for missing."""
    if not rows:
        return []
    headers = list(rows[0].keys())
    kw_col = find_column(headers, COLUMN_ALIASES["keyword"])
    if kw_col is None:
        raise ValueError(
            f"Geen keyword-kolom gevonden. Gevonden kolommen: {list(rows[0].keys())}. "
            "Verwacht: keyword, query, search term, zoekwoord, of topic."
        )
    vol_col = find_column(headers, COLUMN_ALIASES["volume"])
    kd_col = find_column(headers, COLUMN_ALIASES["kd"])
    pos_col = find_column(headers, COLUMN_ALIASES["position"])
    cpc_col = find_column(headers, COLUMN_ALIASES["cpc"])
    url_col = find_column(headers, COLUMN_ALIASES["url"])
    cluster_col = find_column(headers, COLUMN_ALIASES["cluster"])
    cp_col = find_column(headers, COLUMN_ALIASES["click_potential"])
    serp_col = find_column(headers, COLUMN_ALIASES["serp_features_raw"])
    intent_col = find_column(headers, COLUMN_ALIASES["intent"])

    mkt = (market or "NL").strip().upper()
    if mkt not in ("NL", "UK", "DE"):
        mkt = "NL"

    result = []
    for r in rows:
        kw = (r.get(kw_col) or "").strip()
        if not kw:
            continue
        vol_raw = r.get(vol_col) if vol_col else None
        kd_raw = r.get(kd_col) if kd_col else None
        try:
            vol = int(float(str(vol_raw or "0").replace(",", "").replace(" ", "")))
        except (ValueError, TypeError):
            vol = 0
        try:
            kd = float(str(kd_raw or "50").replace(",", ".").replace(" ", ""))
        except (ValueError, TypeError):
            kd = 50.0

        pos = 0
        if pos_col and r.get(pos_col):
            try:
                pos = int(float(str(r.get(pos_col)).replace(",", "")))
            except (ValueError, TypeError):
                pass

        cpc = 0.0
        if cpc_col and r.get(cpc_col):
            try:
                cpc = float(str(r.get(cpc_col)).replace(",", "."))
            except (ValueError, TypeError):
                pass

        url = (r.get(url_col) or "").strip() if url_col else ""
        intent = (r.get(intent_col) or "").strip() if intent_col else ""
        cluster = (r.get(cluster_col) or "").strip() if cluster_col else ""
        if cluster.lower() in ("nan", "#n/a", "-", "none"):
            cluster = ""
        serp_raw = (r.get(serp_col) or "").strip() if serp_col else ""
        cp_val = _parse_optional_float(r.get(cp_col)) if cp_col else None

        result.append({
            "keyword": kw,
            "volume": vol,
            "kd": kd,
            "position": pos,
            "cpc": cpc,
            "current_url": url or None,
            "intent": intent or None,
            "cluster": cluster or None,
            "click_potential": cp_val,
            "serp_features_raw": serp_raw or None,
            "market": mkt,
        })
    return result


def parse_keywords_file(content: bytes, filename: str, market: str = "NL") -> List[Dict[str, Any]]:
    """
    Parse CSV, XLSX, or Numbers file into keyword list.
    Only a recognizable keyword column is required; other columns use defaults if missing.
    ``market`` tags rows as NL / UK / DE for combined plans.
    Raises ValueError on no keyword column or invalid format.
    """
    ext = (filename or "").lower().split(".")[-1] if "." in (filename or "") else ""
    if ext == "csv":
        return _normalize_rows(parse_csv(content), market=market)
    if ext in ("xlsx", "xls"):
        return _normalize_rows(parse_xlsx(content), market=market)
    if ext == "numbers":
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(suffix=".numbers", delete=False) as tmp:
                tmp.write(content)
                tmp_path = tmp.name
            data = parse_numbers_file(tmp_path)
            return _normalize_rows(data, market=market)
        finally:
            if tmp_path and os.path.isfile(tmp_path):
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass
    raise ValueError(f"Unsupported file type: .{ext}. Use CSV, XLSX, or Numbers.")


def load_keywords_job_file(path: str) -> List[Dict[str, Any]]:
    """
    Load keywords from a job input path: merged JSON (multi-market) or CSV/XLSX/Numbers.
    """
    lower = path.lower()
    if lower.endswith(".json"):
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, list):
            raise ValueError("Keywords JSON must be a list")
        return data
    with open(path, "rb") as f:
        content = f.read()
    return parse_keywords_file(content, os.path.basename(path), market="NL")
