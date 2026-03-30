#!/usr/bin/env python3
"""
SEO Tool Output Validator
=========================
Valideert de output van de SEO tool (Excel of CSV) tegen 3 structurele criteria
+ uitgebreide checks op basis van de SEO playbook.

Gebruik:
    python3 seo_validator.py <output.xlsx>
    python3 seo_validator.py <output.xlsx> --csv-source <source.csv>
    python3 seo_validator.py <output.xlsx> --verbose
    python3 seo_validator.py <output.xlsx> --report

Output:
    - Console: PASS/FAIL per check met details
    - File: <output_stem>_validation_report.md (Markdown rapport; met --report)

Exit codes: 0 = clean, 1 = warnings, 2 = errors (CI: fail op 2)
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd

# ── Config ────────────────────────────────────────────────────────────────────
VETO_KD_THRESHOLD = 45  # KD boven dit = nooit HIGH zonder ranking
SILO_CONCENTRATION_MAX = 0.40  # geen silo mag >40% van alle keywords bevatten
QUICK_WIN_MIN_VOLUME = 100
QUICK_WIN_MAX_KD = 40
REQUIRED_TABS = {
    "keyword plan",
    "silo",
    "quick win",
    "content gap",
    "strategie",
    "gsc",
}
BANNED_SILO_NAMES = {"overig", "algemeen", "other", "misc", "rest", "undefined"}
REQUIRED_COLUMNS = [
    "keyword",
    "silo",
    "volume",
    "kd",
    "intent",
    "status (gsc)",
    "prioriteit",
    "content type",
]


# ── Helpers ────────────────────────────────────────────────────────────────────
def normalize(s):
    return str(s).lower().strip() if pd.notna(s) else ""


def find_sheet(wb, keywords):
    """Find sheet by partial name match (case-insensitive)."""
    for name in wb.sheetnames:
        if any(kw in name.lower() for kw in keywords):
            return wb[name]
    return None


def _pad_row(row, n: int) -> list:
    row = list(row)
    if len(row) < n:
        row.extend([None] * (n - len(row)))
    return row[:n]


def read_keyword_sheet(wb):
    """Read the keyword plan sheet into a DataFrame."""
    sheet = find_sheet(wb, ["keyword", "plan", "zoekwoord"])
    if not sheet:
        return None, "Geen Keyword Plan sheet gevonden"

    data = list(sheet.values)
    # Find the actual header row (skip title rows)
    hdr_idx = None
    for i, row in enumerate(data):
        vals = [normalize(c) for c in row if c is not None and str(c).strip() != ""]
        if "keyword" in vals and "prioriteit" in vals:
            hdr_idx = i
            break
    if hdr_idx is None:
        return None, "Header rij niet gevonden in Keyword Plan sheet"

    headers = [normalize(c) for c in data[hdr_idx]]
    ncols = len(headers)
    rows = []
    for row in data[hdr_idx + 1 :]:
        row = _pad_row(row, ncols)
        if any(c is not None and str(c).strip() != "" for c in row):
            rows.append(dict(zip(headers, row)))

    df = pd.DataFrame(rows)
    return df, None


# ── Result collector ───────────────────────────────────────────────────────────
class ValidationResult:
    def __init__(self):
        self.checks = []

    def add(self, name, passed, details="", severity="ERROR", data=None):
        self.checks.append(
            {
                "name": name,
                "passed": passed,
                "details": details,
                "severity": severity,  # ERROR | WARNING | INFO
                "data": data or [],
            }
        )

    def passed_count(self):
        return sum(1 for c in self.checks if c["passed"])

    def failed_errors(self):
        return [c for c in self.checks if not c["passed"] and c["severity"] == "ERROR"]

    def failed_warnings(self):
        return [c for c in self.checks if not c["passed"] and c["severity"] == "WARNING"]

    def score(self):
        total = len(self.checks)
        return round(self.passed_count() / total * 100) if total else 0


# ── CHECKS ─────────────────────────────────────────────────────────────────────


def check_required_tabs(wb, r):
    present = [name.lower() for name in wb.sheetnames]
    found = set()
    for req in REQUIRED_TABS:
        if any(req in p for p in present):
            found.add(req)
    missing = REQUIRED_TABS - found
    r.add(
        "Verplichte tabs aanwezig",
        len(missing) == 0,
        f"Aanwezig: {sorted(found)} | Ontbreekt: {sorted(missing)}",
        severity="ERROR",
    )


def check_required_columns(df, r):
    if df is None:
        r.add("Verplichte kolommen", False, "Kan sheet niet lezen", severity="ERROR")
        return
    cols = [normalize(c) for c in df.columns]
    missing = [req for req in REQUIRED_COLUMNS if req not in cols]
    r.add(
        "Verplichte kolommen aanwezig",
        len(missing) == 0,
        f"Ontbrekende kolommen: {missing}" if missing else "Alle verplichte kolommen aanwezig",
        severity="ERROR",
    )


def check_silo_concentration(df, r):
    """Criterion 1: geen silo >40% van keywords."""
    if df is None or "silo" not in df.columns:
        r.add("Silo concentratie", False, "Silo kolom ontbreekt", severity="ERROR")
        return

    silo_counts = df["silo"].apply(normalize).value_counts()
    total = len(df)
    offenders = {
        silo: count
        for silo, count in silo_counts.items()
        if count / total > SILO_CONCENTRATION_MAX and silo
    }

    details_lines = []
    for silo, count in silo_counts.items():
        pct = count / total * 100
        flag = " ← ❌ te groot" if silo in offenders else ""
        details_lines.append(f"  {silo or '(leeg)'}: {count} keywords ({pct:.1f}%){flag}")

    r.add(
        f"Silo concentratie ≤{int(SILO_CONCENTRATION_MAX * 100)}% per silo",
        len(offenders) == 0,
        "\n".join(details_lines),
        severity="ERROR",
        data=list(offenders.keys()),
    )


def check_no_vangnet_silo(df, r):
    """Geen vangnet-namen zoals 'Overig'."""
    if df is None or "silo" not in df.columns:
        r.add("Geen vangnet-silos", False, "Silo kolom ontbreekt", severity="ERROR")
        return

    silos_present = {normalize(s) for s in df["silo"].dropna().unique()}
    # Exact match only - 'Overige Commercial' is OK, standalone 'Overig' is not
    found_banned = {s for s in silos_present if s in BANNED_SILO_NAMES}
    # Partial: only flag if the silo name IS a banned word (not just contains it)
    single_word_bad = {
        s
        for s in silos_present
        if s.split() and s.split()[0] in BANNED_SILO_NAMES and len(s.split()) == 1
    }
    all_bad = found_banned | single_word_bad

    r.add(
        "Geen vangnet-silos (Overig / Misc / etc.)",
        len(all_bad) == 0,
        f"Gevonden vangnet-silos: {all_bad}" if all_bad else "Geen vangnet-silos gevonden",
        severity="ERROR",
        data=list(all_bad),
    )


def check_priority_veto(df, r):
    """Criterion 2: KD >45 zonder GSC ranking mag nooit HIGH zijn."""
    if df is None:
        r.add("Prioriteit veto (KD>45, geen ranking)", False, "Sheet niet beschikbaar", severity="ERROR")
        return

    if "kd" not in df.columns or "prioriteit" not in df.columns:
        r.add(
            "Prioriteit veto (KD>45, geen ranking)",
            False,
            "KD of Prioriteit kolom ontbreekt",
            severity="ERROR",
        )
        return

    status_col = next((c for c in df.columns if "status" in c and "gsc" in c), None)

    violations = []
    for _, row in df.iterrows():
        try:
            kd = float(row.get("kd") or 0)
        except (ValueError, TypeError):
            continue

        prio = normalize(row.get("prioriteit", ""))
        status = normalize(row.get(status_col, "") if status_col else "")

        no_ranking = "⬜" in status or "ontbreekt" in status or not status
        if kd > VETO_KD_THRESHOLD and no_ranking and prio == "high":
            kw = row.get("keyword", "?")
            violations.append(f"  '{kw}' — KD: {kd:.0f}, Status: {status or 'leeg'}, Prioriteit: HIGH")

    r.add(
        f"Prioriteit veto: KD>{VETO_KD_THRESHOLD} zonder ranking ≠ HIGH",
        len(violations) == 0,
        f"{len(violations)} violations:\n" + "\n".join(violations[:10]) if violations else "Geen violations",
        severity="ERROR",
        data=violations,
    )


def check_specific_keywords(df, r):
    """Criterion 3: spot-checks op bekende keywords."""
    if df is None:
        r.add("Spot-checks specifieke keywords", False, "Sheet niet beschikbaar", severity="ERROR")
        return
    if "keyword" not in df.columns:
        r.add("Spot-checks specifieke keywords", False, "Keyword kolom ontbreekt", severity="ERROR")
        return

    status_col = next((c for c in df.columns if "status" in c and "gsc" in c), None)
    clicks_col = next((c for c in df.columns if "click" in c and "gsc" in c), None)

    spot_checks = []

    # Check 1: reisverzekering moet MEDIUM zijn (KD 63, geen ranking)
    row_rv = df[df["keyword"].apply(normalize) == "reisverzekering"]
    if not row_rv.empty:
        prio = normalize(row_rv.iloc[0].get("prioriteit", ""))
        kd_val = row_rv.iloc[0].get("kd", "?")
        if prio == "high":
            spot_checks.append(f"  ❌ 'reisverzekering' (KD={kd_val}): staat als HIGH, verwacht MEDIUM")
        else:
            spot_checks.append(f"  ✅ 'reisverzekering' (KD={kd_val}): correct als {prio.upper()}")
    else:
        spot_checks.append("  ⚠️  'reisverzekering' niet gevonden in keyword plan")

    # Check 2: allianz huurauto verzekering moet hoge QW score hebben
    row_allianz = df[df["keyword"].apply(normalize).str.contains("allianz huurauto", na=False)]
    if not row_allianz.empty and clicks_col:
        clicks = row_allianz.iloc[0].get(clicks_col, 0)
        try:
            clicks_int = int(float(clicks)) if pd.notna(clicks) else 0
        except (ValueError, TypeError):
            clicks_int = 0

        if clicks_int >= 100:
            spot_checks.append(
                f"  ✅ 'allianz huurauto verzekering': {clicks_int} GSC clicks — verwacht hoge QW score"
            )
        else:
            spot_checks.append(
                f"  ⚠️  'allianz huurauto verzekering': {clicks_int} GSC clicks — check QW score"
            )
    else:
        spot_checks.append("  ⚠️  'allianz huurauto verzekering' niet gevonden of geen clicks kolom")

    # Check 3: winterbanden verplicht duitsland — moet MEDIUM of HIGH zijn (vol 1600, KD 14)
    row_wb = df[df["keyword"].apply(normalize).str.contains("winterbanden verplicht", na=False)]
    if not row_wb.empty:
        prio = normalize(row_wb.iloc[0].get("prioriteit", ""))
        vol = row_wb.iloc[0].get("volume", "?")
        if prio == "low":
            spot_checks.append(
                f"  ❌ 'winterbanden verplicht duitsland' (vol={vol}): staat als LOW, verwacht MEDIUM/HIGH"
            )
        else:
            spot_checks.append(f"  ✅ 'winterbanden verplicht duitsland' (vol={vol}): correct als {prio.upper()}")
    else:
        spot_checks.append("  ⚠️  'winterbanden verplicht duitsland' niet gevonden")

    r.add(
        "Spot-checks specifieke keywords",
        not any("❌" in s for s in spot_checks),
        "\n".join(spot_checks),
        severity="ERROR",
        data=spot_checks,
    )


def check_quick_win_filter(wb, r):
    """Quick Wins sheet: filter criteria correct toegepast."""
    sheet = find_sheet(wb, ["quick win", "quick", "win"])
    if not sheet:
        r.add("Quick Wins filter", False, "Quick Wins sheet niet gevonden", severity="WARNING")
        return

    data = list(sheet.values)
    hdr_idx = None
    for i, row in enumerate(data):
        vals = [normalize(c) for c in row if c is not None and str(c).strip() != ""]
        # Real header must have BOTH keyword-like AND at least one metric column
        has_kw = any(v in ("keyword", "zoekwoord") for v in vals)
        has_metric = any(v in ("kd", "volume", "prioriteit", "status (gsc)") for v in vals)
        if has_kw and has_metric:
            hdr_idx = i
            break
    if hdr_idx is None:
        r.add("Quick Wins filter", False, "Header niet gevonden in QW sheet", severity="WARNING")
        return

    headers = [normalize(c) for c in data[hdr_idx]]
    ncols = len(headers)
    rows = []
    for row in data[hdr_idx + 1 :]:
        row = _pad_row(row, ncols)
        if any(c is not None and str(c).strip() != "" for c in row):
            rows.append(dict(zip(headers, row)))
    df_qw = pd.DataFrame(rows)
    # Skip rows with empty keyword
    kw_col_qw = next((c for c in df_qw.columns if "keyword" in c or "zoekwoord" in c), None)
    if kw_col_qw:
        df_qw = df_qw[df_qw[kw_col_qw].apply(lambda x: bool(normalize(x)) and normalize(x) != "nan")]
    else:
        df_qw = df_qw[df_qw.apply(lambda row_: any(normalize(v) for v in row_.values), axis=1)]

    violations = []
    for _, row in df_qw.iterrows():
        try:
            vol = float(row.get("volume") or 0)
            kd = float(row.get("kd") or 99)
        except (ValueError, TypeError):
            continue
        if not vol and not kd:
            continue  # skip completely empty data rows

        status_key = next((k for k in row.keys() if "status" in k and "gsc" in k), None)
        status = normalize(row.get(status_key, "") if status_key else "")
        kw = row.get(kw_col_qw or "keyword", "?")

        if vol < QUICK_WIN_MIN_VOLUME:
            violations.append(f"  '{kw}': volume {vol:.0f} < {QUICK_WIN_MIN_VOLUME}")
        if kd > QUICK_WIN_MAX_KD:
            violations.append(f"  '{kw}': KD {kd:.0f} > {QUICK_WIN_MAX_KD}")
        if status and not any(s in status for s in ["aanpakken", "optimaliseer", "🟠", "🟡"]):
            violations.append(f"  '{kw}': status '{status}' voldoet niet aan filter")

    r.add(
        f"Quick Wins filter (vol≥{QUICK_WIN_MIN_VOLUME}, KD≤{QUICK_WIN_MAX_KD}, status 🟠/🟡)",
        len(violations) == 0,
        f"{len(violations)} filter-violations:\n" + "\n".join(violations[:10])
        if violations
        else f"{len(df_qw)} Quick Wins — alle voldoen aan filter",
        severity="WARNING",
        data=violations,
    )


def check_gsc_data_present(df, r):
    """GSC data is aanwezig voor rankings (niet alles leeg)."""
    if df is None:
        r.add("GSC data aanwezig", False, "Sheet niet beschikbaar", severity="ERROR")
        return

    status_col = next((c for c in df.columns if "status" in c and "gsc" in c), None)
    clicks_col = next((c for c in df.columns if "click" in c and "gsc" in c), None)

    if not status_col:
        r.add("GSC data aanwezig", False, "Geen GSC Status kolom gevonden", severity="ERROR")
        return

    total = len(df)
    with_gsc = df[df[status_col].apply(normalize) != "⬜ ontbreekt"].shape[0]
    pct = with_gsc / total * 100 if total else 0

    # Check that keywords with clicks actually have them filled in
    clicks_filled = 0
    if clicks_col:
        clicks_filled = df[
            df[clicks_col].apply(lambda x: pd.notna(x) and str(x).strip() not in ("", "0", "None", "nan"))
        ].shape[0]

    details = (
        f"Keywords met GSC-status: {with_gsc}/{total} ({pct:.0f}%)\n"
        f"Keywords met clicks ingevuld: {clicks_filled}"
    )

    # At least 10% should have GSC data for a live site
    r.add(
        "GSC data aanwezig (>10% keywords hebben GSC-status)",
        pct >= 10,
        details,
        severity="ERROR",
    )


def check_serp_features_present(df, r):
    """SERP Features kolom bestaat en is gevuld."""
    if df is None:
        r.add("SERP Features kolom", False, "Sheet niet beschikbaar", severity="WARNING")
        return

    serp_col = next((c for c in df.columns if "serp" in c), None)
    if not serp_col:
        r.add("SERP Features kolom", False, "SERP Features kolom ontbreekt volledig", severity="ERROR")
        return

    filled = df[
        df[serp_col].apply(lambda x: pd.notna(x) and str(x).strip() not in ("", "nan"))
    ].shape[0]
    pct = filled / len(df) * 100 if len(df) else 0

    r.add(
        "SERP Features kolom gevuld (≥50% ingevuld)",
        pct >= 50,
        f"Gevuld: {filled}/{len(df)} ({pct:.0f}%)",
        severity="WARNING",
    )


def check_doelgroep_diversity(df, r):
    """Doelgroep is niet allemaal hetzelfde."""
    if df is None or "doelgroep" not in df.columns:
        r.add("Doelgroep diversiteit", False, "Doelgroep kolom ontbreekt", severity="WARNING")
        return

    unique_dg = df["doelgroep"].apply(normalize).nunique()
    top = df["doelgroep"].apply(normalize).value_counts().iloc[0] if len(df) else 0
    top_pct = top / len(df) * 100 if len(df) else 0
    top_name = df["doelgroep"].apply(normalize).value_counts().index[0] if len(df) else ""

    r.add(
        "Doelgroep gedifferentieerd (niet allemaal hetzelfde)",
        unique_dg >= 3 and top_pct < 70,
        f"Unieke doelgroepen: {unique_dg} | Grootste: '{top_name}' ({top_pct:.0f}%)",
        severity="WARNING",
    )


def check_markt_column(df, r):
    """Markt kolom aanwezig met NL/UK/DE waarden."""
    if df is None:
        r.add("Markt kolom", False, "Sheet niet beschikbaar", severity="WARNING")
        return

    if "markt" not in df.columns:
        r.add("Markt kolom aanwezig", False, "Markt kolom ontbreekt", severity="WARNING")
        return

    markets = df["markt"].apply(normalize).value_counts().to_dict()
    r.add(
        "Markt kolom aanwezig",
        True,
        f"Verdeling: {markets}",
        severity="INFO",
    )


def check_content_gaps_priority_cap(wb, r):
    """Content Gaps: KD>45 zonder ranking = max MEDIUM."""
    sheet = find_sheet(wb, ["content gap", "gap", "ontbreekt"])
    if not sheet:
        r.add("Content Gaps priority cap", False, "Content Gaps sheet niet gevonden", severity="WARNING")
        return

    data = list(sheet.values)
    hdr_idx = None
    for i, row in enumerate(data):
        vals = [normalize(c) for c in row if c is not None and str(c).strip() != ""]
        has_kw = any(v in ("keyword", "zoekwoord") for v in vals)
        has_metric = any(v in ("kd", "prioriteit", "volume", "intent") for v in vals)
        if has_kw and has_metric:
            hdr_idx = i
            break
    if hdr_idx is None:
        r.add("Content Gaps priority cap", False, "Header niet gevonden", severity="WARNING")
        return

    headers = [normalize(c) for c in data[hdr_idx]]
    ncols = len(headers)
    rows = []
    for row in data[hdr_idx + 1 :]:
        row = _pad_row(row, ncols)
        if any(c is not None and str(c).strip() != "" for c in row):
            rows.append(dict(zip(headers, row)))
    df_gaps = pd.DataFrame(rows)

    kd_col = next((c for c in df_gaps.columns if c in ("kd", "keyword difficulty", "moeilijkheid")), None)
    prio_col_g = next((c for c in df_gaps.columns if "prioriteit" in c or "priority" in c), None)
    if not kd_col or not prio_col_g:
        r.add(
            "Content Gaps priority cap",
            False,
            f"KD ({kd_col}) of Prioriteit ({prio_col_g}) kolom ontbreekt. Gevonden: {list(df_gaps.columns[:10])}",
            severity="WARNING",
        )
        return

    violations = []
    for _, row in df_gaps.iterrows():
        try:
            kd = float(row.get(kd_col) or 0)
        except (ValueError, TypeError):
            continue
        prio = normalize(row.get(prio_col_g, ""))
        kw = row.get("keyword", "?")
        if not kw or normalize(kw) == "nan":
            continue
        if kd > VETO_KD_THRESHOLD and prio == "high":
            violations.append(f"  '{kw}' — KD: {kd:.0f} → HIGH (max MEDIUM)")

    r.add(
        f"Content Gaps priority cap (KD>{VETO_KD_THRESHOLD} = max MEDIUM)",
        len(violations) == 0,
        f"{len(violations)} violations:\n" + "\n".join(violations[:10])
        if violations
        else f"Geen violations in {len(df_gaps)} Content Gaps",
        severity="ERROR",
        data=violations,
    )


def check_no_empty_priorities(df, r):
    """Geen lege Prioriteit waarden."""
    if df is None or "prioriteit" not in df.columns:
        r.add("Geen lege prioriteiten", False, "Prioriteit kolom ontbreekt", severity="WARNING")
        return

    empty = df[df["prioriteit"].apply(lambda x: not normalize(x) or normalize(x) == "nan")].shape[0]
    r.add(
        "Geen lege Prioriteit waarden",
        empty == 0,
        f"{empty} rijen zonder prioriteit" if empty else "Alle rijen hebben een prioriteit",
        severity="WARNING",
    )


# ── MAIN ───────────────────────────────────────────────────────────────────────


def run_validation(xlsx_path: str, csv_source: Optional[str] = None) -> ValidationResult:
    r = ValidationResult()

    print(f"\n{'=' * 60}")
    print(f"  SEO Tool Validator — {Path(xlsx_path).name}")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"{'=' * 60}\n")

    if csv_source:
        p = Path(csv_source)
        if not p.exists():
            print(f"⚠️  --csv-source bestand niet gevonden: {csv_source} (genegeerd)\n")
        else:
            print(f"ℹ️  --csv-source {csv_source}: optionele cross-checks volgen in een latere versie.\n")

    # Load workbook
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        print(f"❌ FATAL: Kan Excel niet openen: {e}")
        sys.exit(1)

    # Load keyword plan DF
    df, err = read_keyword_sheet(wb)
    if err:
        print(f"⚠️  Waarschuwing: {err}")

    # Run all checks
    check_required_tabs(wb, r)
    check_required_columns(df, r)
    check_silo_concentration(df, r)  # Criterion 1
    check_no_vangnet_silo(df, r)
    check_priority_veto(df, r)  # Criterion 2
    check_specific_keywords(df, r)  # Criterion 3
    check_quick_win_filter(wb, r)
    check_gsc_data_present(df, r)
    check_serp_features_present(df, r)
    check_doelgroep_diversity(df, r)
    check_markt_column(df, r)
    check_content_gaps_priority_cap(wb, r)
    check_no_empty_priorities(df, r)

    return r


def print_results(r: ValidationResult, verbose: bool = False):
    errors = r.failed_errors()
    warnings = r.failed_warnings()
    score = r.score()

    # Summary bar
    bar_len = 40
    filled = int(bar_len * score / 100)
    bar = "█" * filled + "░" * (bar_len - filled)
    color = "✅" if score >= 80 else ("⚠️ " if score >= 60 else "❌")

    print(f"Score: {color} {score}%  [{bar}]")
    print(f"       {r.passed_count()}/{len(r.checks)} checks geslaagd\n")

    # Errors
    if errors:
        print(f"{'─' * 50}")
        print(f"  ERRORS ({len(errors)}) — blokkeren correcte output")
        print(f"{'─' * 50}")
        for c in errors:
            print(f"\n  ❌ {c['name']}")
            if verbose or len(c["details"]) < 200:
                for line in c["details"].split("\n"):
                    print(f"     {line}")
            else:
                preview = c["details"][:200].split("\n")[0]
                print(f"     {preview}... (gebruik --verbose voor details)")

    # Warnings
    if warnings:
        print(f"\n{'─' * 50}")
        print(f"  WARNINGS ({len(warnings)}) — verbetering aanbevolen")
        print(f"{'─' * 50}")
        for c in warnings:
            print(f"\n  ⚠️  {c['name']}")
            if verbose:
                for line in c["details"].split("\n"):
                    print(f"     {line}")
            else:
                print(f"     {c['details'].split(chr(10))[0][:120]}")

    # Passed
    passed = [c for c in r.checks if c["passed"]]
    print(f"\n{'─' * 50}")
    print(f"  GESLAAGD ({len(passed)})")
    print(f"{'─' * 50}")
    for c in passed:
        print(f"  ✅ {c['name']}")
        if verbose and c["details"]:
            for line in c["details"].split("\n")[:3]:
                print(f"     {line}")

    print(f"\n{'=' * 60}\n")

    # Verdict
    if not errors and not warnings:
        print("🏆 Perfecte score — output voldoet aan alle criteria.")
    elif not errors:
        print("✅ Geen blokkerende errors. Warnings zijn verbeterpunten.")
    else:
        print(f"❌ {len(errors)} error(s) gevonden. Fix deze vóór gebruik in productie.")

    print()


def generate_report(r: ValidationResult, xlsx_path: str) -> str:
    """Generate a Markdown report."""
    score = r.score()
    errors = r.failed_errors()
    warnings = r.failed_warnings()
    passed = [c for c in r.checks if c["passed"]]

    lines = [
        "# SEO Tool Validatierapport",
        f"**Bestand:** `{Path(xlsx_path).name}`  ",
        f"**Datum:** {datetime.now().strftime('%Y-%m-%d %H:%M')}  ",
        f"**Score:** {score}% ({r.passed_count()}/{len(r.checks)} checks geslaagd)  ",
        "",
    ]

    if errors:
        lines += [f"## ❌ Errors ({len(errors)})", ""]
        for c in errors:
            lines.append(f"### {c['name']}")
            lines.append("```")
            lines.append(c["details"][:800])
            lines.append("```")
            lines.append("")

    if warnings:
        lines += [f"## ⚠️ Warnings ({len(warnings)})", ""]
        for c in warnings:
            lines.append(f"### {c['name']}")
            lines.append(c["details"][:400])
            lines.append("")

    if passed:
        lines += [f"## ✅ Geslaagd ({len(passed)})", ""]
        for c in passed:
            lines.append(f"- **{c['name']}**: {c['details'].split(chr(10))[0][:120]}")
        lines.append("")

    lines += [
        "## Aanbevolen acties",
        "",
        "Prioriteer in deze volgorde:",
    ]
    for i, c in enumerate(errors, 1):
        lines.append(f"{i}. **[ERROR]** {c['name']}")
    for j, c in enumerate(warnings, len(errors) + 1):
        lines.append(f"{j}. **[WARNING]** {c['name']}")

    return "\n".join(lines)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="SEO Tool Output Validator")
    parser.add_argument("xlsx", help="Pad naar SEO output Excel file")
    parser.add_argument("--verbose", "-v", action="store_true", help="Toon volledige details")
    parser.add_argument("--report", "-r", action="store_true", help="Sla Markdown rapport op")
    parser.add_argument(
        "--csv-source",
        default=None,
        help="(optioneel) Bron-CSV voor toekomstige cross-checks; nu alleen gecontroleerd op bestaan",
    )
    args = parser.parse_args()

    if not Path(args.xlsx).exists():
        print(f"❌ Bestand niet gevonden: {args.xlsx}")
        sys.exit(1)

    results = run_validation(args.xlsx, csv_source=args.csv_source)
    print_results(results, verbose=args.verbose)

    if args.report:
        xlsx_p = Path(args.xlsx).expanduser().absolute()
        report_path = xlsx_p.parent / f"{xlsx_p.stem}_validation_report.md"
        with open(report_path, "w", encoding="utf-8") as f:
            f.write(generate_report(results, args.xlsx))
        print(f"📄 Rapport opgeslagen: {report_path}")

    # Exit code: 0 = all pass, 1 = warnings, 2 = errors
    if results.failed_errors():
        sys.exit(2)
    elif results.failed_warnings():
        sys.exit(1)
    else:
        sys.exit(0)
